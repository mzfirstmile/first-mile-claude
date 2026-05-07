#!/usr/bin/env python3
"""
inspect_rent_roll.py — Dump the first ~40 rows of a rent-roll xlsx so we can
figure out where the header row actually lives. Use this when sync_rent_rolls.py
fails with "Could not locate header row".

Usage:
    python3 scripts/inspect_rent_roll.py <path-to-xlsx>
    python3 scripts/inspect_rent_roll.py --property "61 S Paramus"
"""

from __future__ import annotations

import argparse
import glob
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


DROPBOX_ROOTS = [
    Path.home() / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "Dropbox-FirstMileCapital",
    Path.home() / "First Mile Dropbox" / "Morris Zeitouni" / "First Mile Prop Dropbox",
    Path.home() / "First Mile Dropbox" / "First Mile Prop Dropbox",
    Path.home() / "Dropbox" / "First Mile Prop Dropbox",
    Path.home() / "First Mile Dropbox",
]
PROPMGMT_DIRNAME = "2.1 FMC Property Management"
PREFERRED_SUBDIRS = [
    "3 - Operations", "2 - Leasing_Marketing", "2 - Leasing", "4 - Accounting",
]


def find_dropbox_root() -> Path | None:
    for p in DROPBOX_ROOTS:
        if p.exists() and p.is_dir():
            return p
    return None


def find_rent_roll_in_property(prop_substr: str) -> Path | None:
    root = find_dropbox_root()
    if not root:
        sys.exit("Dropbox root not found")
    propmgmt = root / PROPMGMT_DIRNAME
    if not propmgmt.exists():
        sys.exit(f"{PROPMGMT_DIRNAME} not found under {root}")

    matches = [c for c in propmgmt.iterdir() if c.is_dir() and prop_substr.lower() in c.name.lower()]
    if not matches:
        sys.exit(f"No property folder matching '{prop_substr}'")
    if len(matches) > 1:
        print("Multiple matches:", [m.name for m in matches])
    prop_folder = matches[0]
    print(f"✓ Property folder: {prop_folder}")

    patterns = [
        "Rent Roll*.xlsx", "RentRoll*.xlsx", "rent_roll*.xlsx", "RR*.xlsx",
        "*rent roll*.xlsx", "*rentroll*.xlsx",
    ]
    candidates: list[Path] = []
    for sub in PREFERRED_SUBDIRS:
        sub_path = prop_folder / sub
        if not sub_path.exists():
            continue
        for pat in patterns:
            candidates.extend(Path(p) for p in glob.glob(str(sub_path / pat)))
            candidates.extend(Path(p) for p in glob.glob(str(sub_path / "**" / pat), recursive=True))
    if not candidates:
        for pat in patterns:
            candidates.extend(Path(p) for p in glob.glob(str(prop_folder / "**" / pat), recursive=True))

    candidates = [p for p in candidates if not p.name.startswith("~$")]
    if not candidates:
        sys.exit("No rent roll files found in property folder")
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def materialize(path: Path):
    try:
        if path.stat().st_size > 0:
            return
    except OSError:
        return
    import subprocess
    print(f"… materializing online-only file …")
    subprocess.run(["/bin/cat", str(path)], stdout=subprocess.DEVNULL,
                   stderr=subprocess.DEVNULL, timeout=300, check=False)


def dump_workbook(path: Path, rows_to_show: int = 40):
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n=== File: {path} ===")
    print(f"=== Sheets: {wb.sheetnames} ===\n")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"--- Sheet: {sheet_name}  (max_row={ws.max_row}, max_col={ws.max_column}) ---")
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:rows_to_show]):
            # Render each cell with its column letter
            cells = []
            for j, v in enumerate(row):
                if v is None or str(v).strip() == "":
                    continue
                col_letter = openpyxl.utils.get_column_letter(j + 1)
                txt = str(v).strip()
                if len(txt) > 35:
                    txt = txt[:32] + "..."
                cells.append(f"{col_letter}={txt}")
            print(f"  row {i+1:>3}: {' | '.join(cells) if cells else '(empty)'}")
        print()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file_or_property", nargs="?", help="Path to xlsx OR property name substring")
    ap.add_argument("--property", help="Property folder name substring (alt to positional)")
    ap.add_argument("--rows", type=int, default=40)
    args = ap.parse_args()

    target = args.property or args.file_or_property
    if not target:
        sys.exit("Pass a file path or --property '61 S Paramus'")

    # If it looks like a file path (has slash or ends in .xlsx), treat it as one
    is_file_arg = target.lower().endswith(".xlsx") or "/" in target
    if is_file_arg:
        path = Path(target).expanduser().resolve()
        if not path.exists():
            sys.exit(f"File not found: {path}")
    else:
        path = find_rent_roll_in_property(target)
        print(f"✓ Latest rent roll: {path}\n")

    materialize(path)
    dump_workbook(path, rows_to_show=args.rows)


if __name__ == "__main__":
    sys.exit(main())
