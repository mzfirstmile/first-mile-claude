#!/usr/bin/env python3
"""
inspect_budget.py — Dump the first ~50 rows of the latest Budget_Comparison
xlsx for a given property + year so we can build a parser for it.

Reuses the exact discovery logic from sync_budgets.py so we don't have to
re-type long Dropbox paths.

Usage:
    python3 scripts/inspect_budget.py                       # 61 S Paramus 2025 (default)
    python3 scripts/inspect_budget.py --property "Paramus Plaza"
    python3 scripts/inspect_budget.py --year 2024
    python3 scripts/inspect_budget.py --rows 80
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

import openpyxl  # noqa: E402

from sync_budgets import discover_budget_files  # type: ignore
from sync_actuals import materialize_dropbox_file  # type: ignore


def dump(path: Path, rows: int = 50) -> None:
    print(f"\n=== File: {path}")
    print(f"=== Size: {path.stat().st_size:,} bytes")
    if path.stat().st_size == 0:
        print("… materializing online-only file …")
        if not materialize_dropbox_file(path):
            sys.exit("✗ failed to download from Dropbox")
        print(f"✓ materialized ({path.stat().st_size:,} bytes)")
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"=== Sheets: {wb.sheetnames}\n")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"--- Sheet: {sn}  (max_row={ws.max_row}, max_col={ws.max_column}) ---")
        all_rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(all_rows[:rows]):
            cells = []
            for j, v in enumerate(row):
                if v is None or str(v).strip() == "":
                    continue
                txt = str(v).strip()
                if len(txt) > 32:
                    txt = txt[:29] + "..."
                col = openpyxl.utils.get_column_letter(j + 1)
                cells.append(f"{col}={txt}")
            print(f"  row {i+1:>3}: {' | '.join(cells) if cells else '(empty)'}")
        print()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--property", default="61 S Paramus",
                    help="Folder-name substring (default: '61 S Paramus')")
    ap.add_argument("--year", type=int, default=2025)
    ap.add_argument("--rows", type=int, default=50)
    args = ap.parse_args()

    grouped = discover_budget_files(None, args.property, args.year)
    if not grouped:
        sys.exit(f"No budget files found for property='{args.property}' year={args.year}")
    folder, paths = next(iter(grouped.items()))
    if not paths:
        sys.exit(f"No paths in discovered group {folder}")
    print(f"Inspecting: {folder}")
    dump(paths[0], rows=args.rows)


if __name__ == "__main__":
    main()
